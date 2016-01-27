import logging
import re

from flask import current_app as app
import openpyxl
import sqlalchemy

from ..api import EXTRA_FIELDS
from ..models import User, VirtualGroup, AppContext, Need


logger = logging.getLogger(__name__)

TRUES = ('yes', 'y', 'no', 'n', 'x')
USER_FIELDS = ('username', 'firstname', 'surname', 'mobile_number', 'email')


class MissingFieldError(Exception):
    pass


class InvalidEmailError(Exception):
    pass


class DuplicatedUserError(Exception):
    pass


class UnknownUserError(Exception):
    pass


class UnknownGroupError(Exception):
    pass


class SheetProcessor(object):

    object_class = None
    mandatory_fields = ()

    def __init__(self, sheet):
        self._headers = None
        self.sheet = sheet

    @property
    def headers(self):
        if self._headers is None:
            top = next(self.sheet.rows)
            self._headers = [str(f.value).strip().lower() for f in top]
        return self._headers

    def cleanup(self, item):
        for key in item:
            cleanup_func = getattr(self, 'cleanup_{}'.format(key), None)
            if cleanup_func:
                item[key] = cleanup_func(item[key])
        return item

    @property
    def data(self):
        self.check()
        names = self.headers
        rows = self.sheet.rows
        next(rows)
        for row in rows:
            yield self.cleanup(dict(zip(names,
                                        [str(c.value).strip() for c in row])))

    def check(self):
        for field in self.mandatory_fields:
            if field not in self.headers:
                msg = 'missing field "{}" in "{}" sheet'.format(field,
                                                                self.sheet.title)
                raise MissingFieldError(msg)

    def read(self):
        NotImplementedError()


class GroupProcessor(SheetProcessor):
    object_class = VirtualGroup
    mandatory_fields = ('username',)

    def cleanup_username(self, name):
        if not name:
            raise ValueError('name is required')
        return re.sub('\s', '', str(name).lower())

    @property
    def headers(self):
        headers = super(GroupProcessor, self).headers
        return [h.strip().lower() for h in headers]

    def read(self):
        headers = self.headers

        added = []
        updated = []
        deactivated = []
        memberships = {}

        group_names = set(headers) - set(['username'])

        # deactivated
        for group in VirtualGroup.query.all():
            if group.name not in group_names:
                deactivated.append(group)

        # added
        for group_name in group_names:
            group = VirtualGroup.by_name(group_name)

            if not group:
                group = VirtualGroup(name=group_name)
                added.append(group)
            else:
                updated.append(group)

        # memberships
        for item in self.data:
            user_groups = memberships.setdefault(item['username'], [])
            for group_name in group_names:
                if (item.get(group_name) and
                        item[group_name].strip().lower() in TRUES):
                    user_groups.append(group_name)

        return added, updated, deactivated, memberships


class UserProcessor(SheetProcessor):
    object_class = User
    mandatory_fields = ('username', 'email', 'mobile')

    def cleanup_username(self, name):
        if not name:
            raise ValueError('name is required')
        return re.sub('\s', '', str(name).lower())

    def cleanup_mobile(self, mobile):
        return re.sub('[^+0-9]', '', mobile)

    def cleanup_email(self, email):
        if not email:
            raise ValueError('email is required')
        if '@' not in email:
            raise InvalidEmailError('{} is not a valid email'.format(email))
        return email

    def read(self):
        # deactivated
        deactivated = []
        actives = set([x['username'] for x in self.data])
        for user in User.query.all():
            if user.username not in actives:
                deactivated.append(user)

        allowed_fields = list(USER_FIELDS) + EXTRA_FIELDS

        # created
        added = []
        updated = []
        for item in self.data:
            entity = User.by_username(item['username'])
            if not entity:
                entity = User(username=item['username'],
                              email=item['email'],
                              active=app.config['AUTO_ACTIVATE_NEW_USER'],
                              mobile_number=item['mobile'])
                added.append(entity)
            else:
                updated.append(entity)

            for key in item:
                if key in allowed_fields:
                    setattr(entity, key, item[key])

        return added, updated, deactivated


class ExcelConnector(object):
    def __init__(self, session, user):
        self.session = session
        self.user = user

    def _find_dupe(self, message, users):
        for user in users:
            if user.username in message or user.email in message:
                return user
        return 'another user'

    def process_users(self, sheet):
        p = UserProcessor(sheet)
        added, updated, deactivated = p.read()

        for user in added:
            self.session.add(user)

        for user in deactivated:
            user.active = False

        for user in updated:
            if not user.active:
                user.active = True

        try:
            self.session.commit()
        except sqlalchemy.exc.IntegrityError as err:
            dupe_user = self._find_dupe(err.params, added)
            raise DuplicatedUserError('{} for {}'.format(str(err.orig),
                                                         dupe_user))

        for user in added:
            user.generate_token()

    def process_groups(self, sheet):
        p = GroupProcessor(sheet)
        added, updated, deactivated, memberships = p.read()

        for group in added:
            self.session.add(group)

        for group in deactivated:
            group.active = False

        for group in updated:
            if not group.active:
                group.active = True

        for username, groups in memberships.items():
            user = User.by_username(username)
            if not user:
                msg = ('unable to add {} to '
                       'groups, user unknown').format(username)
                raise UnknownUserError(msg)
            for name in groups:
                group = VirtualGroup.by_name(name)
                if not group:
                    # maybe a new group
                    for new_group in added:
                        if new_group.name == name:
                            group = new_group
                            break
                    else:
                        msg = ('unable to add {} to '
                               'group {}, group unknown').format(username,
                                                                 name)
                        raise UnknownGroupError(msg)
                user.groups.append(group)

        self.session.commit()

    def process_teams(self, sheet):
        pass

    def process(self, fobj):
        try:
            with self.session.no_autoflush:
                wb = openpyxl.load_workbook(fobj.stream, read_only=True,
                                            use_iterators=True,
                                            data_only=True)

                sheets = set(wb.sheetnames)

                # objects
                for sheetname in ('users', 'groups', 'teams'):
                    if sheetname in sheets:
                        f_name = 'process_{}'.format(sheetname)
                        f = getattr(self, f_name)
                        if not f:
                            raise Exception('{} not found'.format(f_name))
                        logger.info('handling {} change'.format(sheetname))
                        f(wb[sheetname])
                        logger.info('done handling {} change'.format(sheetname))
                        sheets.remove(sheetname)

                # permissions
                for context in sheets:
                    logger.info('handling permission change for {}'.format(context))

                self.session.commit()
        except:
            self.session.rollback()
            raise


class ExcelExporter(object):

    HEADER_MAPPING = {'mobile_number': 'mobile'}

    def __init__(self, session, user):
        self.session = session
        self.user = user
        self.user_fields = list(USER_FIELDS) + EXTRA_FIELDS

    def export(self, filename):
        workbook = openpyxl.Workbook(optimized_write=True)
        self.export_users(workbook)
        self.export_groups(workbook)
        self.export_permissions(workbook)
        workbook.save(filename)

    def export_users(self, workbook):
        sheet = workbook.create_sheet(title='users')
        sheet.append([self.HEADER_MAPPING.get(f, f) for f in self.user_fields])
        for user in self.session.query(User).order_by(User.username).all():
            sheet.append([getattr(user, f, '') for f in self.user_fields])

    def export_groups(self, workbook):
        sheet = workbook.create_sheet(title='groups')
        all_groups = self.session.query(VirtualGroup).order_by(VirtualGroup.name).all()

        sheet.append(['username'] + [g.name for g in all_groups])
        for user in self.session.query(User).order_by(User.username).all():
            row = [user.username]
            for group in all_groups:
                row.append('yes' if group in user.groups else None)
            sheet.append(row)

    def export_permissions(self, workbook):
        for ctx in self.session.query(AppContext).all():
            sheet = workbook.create_sheet(title=ctx.name)
            all_needs = self.session.query(Need).filter_by(app_context=ctx).all()

            sheet.append(['username'] + [n.dotted for n in all_needs])
            for user in self.session.query(User).order_by(User.username).all():
                row = [user.username]
                for need in all_needs:
                    row.append('yes' if need in user.permissions else None)
                sheet.append(row)
